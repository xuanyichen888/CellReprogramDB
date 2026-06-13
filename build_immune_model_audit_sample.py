"""
Build a human audit workbook for model-adjudicated immune/T-cell QA decisions.

The workbook samples high-risk automatic decisions and includes all automatic
merge decisions, so the curator can estimate whether model adjudication is safe
before applying it to the database.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

INFILE = Path("qa_outputs/immune_recipe_QA_model_adjudicated.csv")
OUTFILE = Path("qa_outputs/immune_model_auto_decision_audit.xlsx")
RANDOM_SEED = 20260602


def risk_flags(row: pd.Series) -> str:
    flags: list[str] = []
    decision = row.get("model_decision", "")
    if decision == "auto_accept" and row.get("model_confidence", "") != "high":
        flags.append("auto_accept_not_high")
    if decision == "auto_accept" and row.get("issue_flags", ""):
        flags.append("auto_accept_has_issue_flags")
    if decision == "auto_accept" and row.get("tcell_focused", "") == "yes":
        flags.append("auto_accept_tcell")
    if decision == "auto_hide" and row.get("default_visible", "") == "yes":
        flags.append("auto_hide_default_visible")
    if decision == "auto_hide" and row.get("tcell_focused", "") == "yes":
        flags.append("auto_hide_tcell")
    if decision == "auto_hide" and row.get("model_confidence", "") != "high":
        flags.append("auto_hide_not_high")
    if decision == "auto_merge_duplicate" and row.get("tcell_focused", "") == "yes":
        flags.append("auto_merge_tcell")
    if decision == "auto_merge_duplicate" and row.get("model_confidence", "") != "high":
        flags.append("auto_merge_not_high")
    return "; ".join(flags)


def add_risk_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out.insert(0, "audit_verdict", "")
    out.insert(1, "audit_notes", "")
    out.insert(2, "audit_risk_flags", out.apply(risk_flags, axis=1))
    return out


def pick_auto_accept(df: pd.DataFrame, n: int = 20) -> pd.DataFrame:
    accept = df[df["model_decision"] == "auto_accept"].copy()
    priority = accept[
        (accept["model_confidence"] != "high")
        | (accept["issue_flags"] != "")
        | (accept["tcell_focused"] == "yes")
        | (accept["default_visible"] == "yes")
    ].copy()
    priority["_risk_score"] = (
        (priority["model_confidence"] != "high").astype(int) * 10
        + (priority["issue_flags"] != "").astype(int) * 5
        + (priority["tcell_focused"] == "yes").astype(int) * 4
        + priority["review_priority"].astype(int).clip(upper=40)
    )
    selected = priority.sort_values(["_risk_score", "row_index"], ascending=[False, True]).head(n)
    if len(selected) < n:
        rest = accept[~accept["row_index"].isin(selected["row_index"])]
        selected = pd.concat([selected, rest.sample(n=n - len(selected), random_state=RANDOM_SEED)], ignore_index=True)
    return selected.drop(columns=["_risk_score"], errors="ignore")


def pick_auto_hide(df: pd.DataFrame, n: int = 20) -> pd.DataFrame:
    hide = df[df["model_decision"] == "auto_hide"].copy()
    priority = hide[
        (hide["default_visible"] == "yes")
        | (hide["tcell_focused"] == "yes")
        | (hide["model_confidence"] != "high")
    ].copy()
    priority["_risk_score"] = (
        (priority["default_visible"] == "yes").astype(int) * 10
        + (priority["tcell_focused"] == "yes").astype(int) * 6
        + (priority["model_confidence"] != "high").astype(int) * 4
        + priority["review_priority"].astype(int).clip(upper=40)
    )
    selected = priority.sort_values(["_risk_score", "row_index"], ascending=[False, True]).head(n)
    if len(selected) < n:
        rest = hide[~hide["row_index"].isin(selected["row_index"])]
        selected = pd.concat([selected, rest.sample(n=n - len(selected), random_state=RANDOM_SEED)], ignore_index=True)
    return selected.drop(columns=["_risk_score"], errors="ignore")


def preferred_columns(df: pd.DataFrame) -> list[str]:
    first = [
        "audit_verdict",
        "audit_notes",
        "audit_risk_flags",
        "model_decision",
        "model_confidence",
        "manual_recommended",
        "manual_reason",
        "model_rationale",
        "recipe_valid",
        "factor_assessment",
        "cell_assessment",
        "duplicate_assessment",
        "suggested_factors",
        "suggested_factor_type",
        "suggested_species",
        "review_priority",
        "default_visible",
        "tcell_focused",
        "issue_flags",
        "row_index",
        "pmid",
        "year",
        "paper_type",
        "confidence",
        "source_cell",
        "target_cell",
        "source_cell_std",
        "target_cell_std",
        "source_cell_broad",
        "target_cell_broad",
        "factors",
        "factor_type",
        "species",
        "conversion_scope",
        "single_tf_status",
        "is_broad_duplicate",
        "broad_duplicate_group_id",
        "validation_needs_review",
        "validation_action",
        "validation_resolution",
        "title",
        "evidence_sentence",
        "notes",
    ]
    return [c for c in first if c in df.columns] + [c for c in df.columns if c not in first]


def write_sheet(wb: Workbook, name: str, df: pd.DataFrame, header_fill: str) -> None:
    ws = wb.create_sheet(name)
    columns = preferred_columns(df)
    ws.append(columns)
    for _, row in df[columns].iterrows():
        ws.append([row.get(col, "") for col in columns])

    fill = PatternFill("solid", fgColor=header_fill)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    widths = {
        "audit_verdict": 16,
        "audit_notes": 34,
        "audit_risk_flags": 34,
        "model_rationale": 46,
        "manual_reason": 34,
        "source_cell": 26,
        "target_cell": 28,
        "source_cell_std": 25,
        "target_cell_std": 25,
        "factors": 36,
        "title": 42,
        "evidence_sentence": 62,
        "notes": 36,
    }
    for idx, col in enumerate(columns, 1):
        letter = get_column_letter(idx)
        ws.column_dimensions[letter].width = widths.get(col, min(max(len(col) + 2, 10), 22))
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=False)


def write_summary(wb: Workbook, df: pd.DataFrame, accept: pd.DataFrame, hide: pd.DataFrame, merge: pd.DataFrame) -> None:
    ws = wb.active
    ws.title = "Summary"
    rows = [
        ["Metric", "Value"],
        ["All adjudicated immune/T-cell rows", len(df)],
        ["Auto accept", int((df["model_decision"] == "auto_accept").sum())],
        ["Auto hide", int((df["model_decision"] == "auto_hide").sum())],
        ["Auto merge duplicate", int((df["model_decision"] == "auto_merge_duplicate").sum())],
        ["Needs manual", int((df["model_decision"] == "needs_manual").sum())],
        ["Manual recommended", int((df["manual_recommended"] == "yes").sum())],
        ["Auto accept audit sample", len(accept)],
        ["Auto hide audit sample", len(hide)],
        ["Auto merge audit rows", len(merge)],
        ["Sampling seed", RANDOM_SEED],
        ["Sampling rule", "Accept/hide sheets prioritize T-cell, default-visible, non-high confidence, and rows with issue flags; merge sheet includes all auto-merge rows."],
    ]
    for row in rows:
        ws.append(row)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="0F766E")
        cell.font = Font(color="FFFFFF", bold=True)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 110
    ws.freeze_panes = "A2"


def main() -> None:
    df = pd.read_csv(INFILE, dtype=str).fillna("")
    df["review_priority"] = pd.to_numeric(df["review_priority"], errors="coerce").fillna(0).astype(int)
    df["row_index"] = pd.to_numeric(df["row_index"], errors="coerce").fillna(-1).astype(int)

    accept = add_risk_columns(pick_auto_accept(df, 20))
    hide = add_risk_columns(pick_auto_hide(df, 20))
    merge = add_risk_columns(df[df["model_decision"] == "auto_merge_duplicate"].copy())
    manual = add_risk_columns(df[df["manual_recommended"] == "yes"].copy())

    wb = Workbook()
    write_summary(wb, df, accept, hide, merge)
    write_sheet(wb, "Auto_Accept_Audit_20", accept, "1F4E79")
    write_sheet(wb, "Auto_Hide_Audit_20", hide, "7C2D12")
    write_sheet(wb, "Auto_Merge_All_31", merge, "581C87")
    write_sheet(wb, "Manual_Recommended", manual, "334155")

    OUTFILE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTFILE)
    print(f"Wrote {OUTFILE}")
    print(f"Audit rows: accept={len(accept)}, hide={len(hide)}, merge={len(merge)}, manual={len(manual)}")


if __name__ == "__main__":
    main()
