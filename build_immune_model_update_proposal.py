"""
Build a dry-run database update proposal from immune model adjudication.

This script does not edit recipes_master_v2.csv. It converts model decisions
into proposed actions with risk levels so a curator can review what would be
written back before any database mutation happens.
"""

from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from mark_duplicates import split_factors

INFILE = Path("qa_outputs/immune_recipe_QA_model_adjudicated.csv")
CSV_OUT = Path("qa_outputs/immune_model_update_proposal.csv")
XLSX_OUT = Path("qa_outputs/immune_model_update_proposal.xlsx")


def clean(value) -> str:
    return "" if pd.isna(value) else str(value).strip()


def truthy(value) -> bool:
    return clean(value).lower() == "true"


def factor_count(value: str) -> int:
    text = clean(value)
    if text.lower() in {"", "not specified", "unknown", "not specified in text", "none", "n/a"}:
        return 0
    # Paren/slash-aware so "Oct3/4 (Pou5f1)" or "miR (a, b)" count as one factor.
    return len(split_factors(text))


def propose(row: pd.Series) -> dict:
    decision = clean(row.get("model_decision"))
    confidence = clean(row.get("model_confidence"))
    manual = clean(row.get("manual_recommended")) == "yes"
    updates: dict[str, str] = {}
    notes: list[str] = []
    actions: list[str] = []
    risk = "low"

    if manual:
        return {
            "proposed_action": "manual_review_required",
            "write_risk": "manual",
            "proposed_updates": "{}",
            "proposal_notes": clean(row.get("manual_reason")) or "Model requested manual review.",
        }

    if decision == "auto_hide":
        current_action = clean(row.get("validation_action"))
        if current_action == "hide_model_adjudicated":
            actions.append("confirm_existing_hide")
            risk = "low"
            notes.append("Already hidden by model adjudication; no database change proposed.")
        elif current_action in {"hide_incomplete_recipe", "remove"}:
            actions.append("confirm_existing_hide")
            updates.update(
                {
                    "validation_needs_review": "False",
                    "validation_action": "hide_model_adjudicated",
                    "validation_resolution": "model_adjudicated_hide",
                    "validation_notes_append": f"Model adjudication ({confidence}): {clean(row.get('model_rationale'))}",
                }
            )
            risk = "medium" if clean(row.get("default_visible")) == "yes" else "low"
            notes.append("Existing hide/remove action confirmed by model.")
        else:
            actions.append("hide_model_adjudicated")
            updates.update(
                {
                    "validation_needs_review": "False",
                    "validation_action": "hide_model_adjudicated",
                    "validation_resolution": "model_adjudicated_hide",
                    "validation_notes_append": f"Model adjudication ({confidence}): {clean(row.get('model_rationale'))}",
                }
            )
            risk = "medium" if clean(row.get("default_visible")) == "yes" else "low"
            notes.append("Hide from default/database view unless curator overrides.")

    elif decision == "auto_merge_duplicate":
        actions.append("confirm_broad_duplicate")
        if not truthy(row.get("is_broad_duplicate")):
            updates["is_broad_duplicate"] = "True"
            updates["broad_duplicate_reason"] = f"model_adjudicated_merge: {clean(row.get('model_rationale'))}"
            risk = "medium"
            notes.append("Would newly mark this row as broad duplicate.")
        else:
            merge_note = f"Model confirmed merge ({confidence}): {clean(row.get('model_rationale'))}"
            if merge_note not in clean(row.get("broad_duplicate_reason")):
                updates["broad_duplicate_reason_append"] = merge_note
            risk = "low"
            notes.append("Already broad duplicate; model confirms merge.")

    elif decision == "auto_accept":
        if clean(row.get("validation_action")) in {"hide_incomplete_recipe", "remove", "fix"}:
            actions.append("accept_cleanup_clear_stale_validation_action")
            updates["validation_action"] = ""
            updates["validation_resolution"] = "model_adjudicated_accept"
            risk = "high"
            notes.append("Clearing an existing hide/fix/remove action can change visibility; review before writing.")
        if truthy(row.get("validation_needs_review")):
            actions.append("accept_cleanup_clear_validation_review")
            updates["validation_needs_review"] = "False"
            updates["validation_resolution"] = updates.get("validation_resolution", "model_adjudicated_accept")
            risk = "medium" if risk != "high" else risk
            notes.append("Model says row is valid; candidate to clear validation review flag.")
        if clean(row.get("single_tf_status")) == "unclear" and factor_count(row.get("factors", "")) == 1:
            actions.append("accept_cleanup_mark_single_tf_valid")
            updates["single_tf_status"] = "standalone_valid"
            updates["validation_resolution"] = updates.get("validation_resolution", "model_adjudicated_single_tf_valid")
            risk = "medium" if risk != "high" else risk
            notes.append("Single-factor recipe judged standalone by model.")
        if truthy(row.get("is_broad_duplicate")):
            actions.append("accept_valid_but_duplicate_hidden")
            notes.append("Recipe appears valid, but broad duplicate flag should remain unless merge review says otherwise.")
        if not actions:
            actions.append("accept_no_change")
            notes.append("Valid row; no database change proposed.")

    else:
        actions.append("manual_review_required")
        risk = "manual"
        notes.append("Unknown or pending model decision.")

    return {
        "proposed_action": "; ".join(actions),
        "write_risk": risk,
        "proposed_updates": json.dumps(updates, ensure_ascii=False, sort_keys=True),
        "proposal_notes": " ".join(notes),
    }


def build_proposal(df: pd.DataFrame) -> pd.DataFrame:
    proposal = df.copy()
    proposed = proposal.apply(propose, axis=1, result_type="expand")
    for col in reversed(["proposal_notes", "proposed_updates", "write_risk", "proposed_action"]):
        proposal.insert(0, col, proposed[col])

    first_cols = [
        "proposed_action",
        "write_risk",
        "proposed_updates",
        "proposal_notes",
        "model_decision",
        "model_confidence",
        "manual_recommended",
        "manual_reason",
        "model_rationale",
        "review_priority",
        "default_visible",
        "tcell_focused",
        "issue_flags",
        "row_index",
        "pmid",
        "year",
        "paper_type",
        "confidence",
        "source_cell",
        "target_cell",
        "source_cell_std",
        "target_cell_std",
        "source_cell_broad",
        "target_cell_broad",
        "factors",
        "factor_type",
        "species",
        "conversion_scope",
        "single_tf_status",
        "is_broad_duplicate",
        "broad_duplicate_reason",
        "broad_duplicate_group_id",
        "validation_needs_review",
        "validation_action",
        "validation_notes",
        "validation_resolution",
        "title",
        "evidence_sentence",
        "notes",
    ]
    cols = [c for c in first_cols if c in proposal.columns] + [c for c in proposal.columns if c not in first_cols]
    return proposal[cols]


def write_sheet(wb: Workbook, name: str, df: pd.DataFrame, header_fill: str) -> None:
    ws = wb.create_sheet(name)
    ws.append(list(df.columns))
    for _, row in df.iterrows():
        ws.append([row.get(col, "") for col in df.columns])
    fill = PatternFill("solid", fgColor=header_fill)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    widths = {
        "proposed_action": 34,
        "proposed_updates": 54,
        "proposal_notes": 54,
        "model_rationale": 48,
        "manual_reason": 34,
        "source_cell": 28,
        "target_cell": 30,
        "factors": 38,
        "title": 46,
        "evidence_sentence": 64,
        "notes": 36,
    }
    for idx, col in enumerate(df.columns, 1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(col, min(max(len(col) + 2, 10), 24))
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=False)


def write_summary(wb: Workbook, proposal: pd.DataFrame) -> None:
    ws = wb.active
    ws.title = "Summary"
    action_counts = proposal["proposed_action"].value_counts()
    risk_counts = proposal["write_risk"].value_counts()
    has_updates = proposal["proposed_updates"].astype(str).ne("{}")
    rows = [
        ["Metric", "Value"],
        ["Total rows", len(proposal)],
        ["Manual review required", int((proposal["proposed_action"] == "manual_review_required").sum())],
        ["Hide proposals", int(proposal["proposed_action"].str.contains("hide|confirm_existing_hide", regex=True).sum())],
        ["Merge proposals", int(proposal["proposed_action"].str.contains("confirm_broad_duplicate").sum())],
        ["Accept/no-change rows", int(proposal["proposed_action"].str.contains("accept_no_change").sum())],
        ["Accept cleanup candidates", int(proposal["proposed_action"].str.contains("accept_cleanup").sum())],
        ["Remaining low-risk write candidates", int(((proposal["write_risk"] == "low") & has_updates).sum())],
        ["Remaining medium-risk write candidates", int(((proposal["write_risk"] == "medium") & has_updates).sum())],
        ["Remaining high-risk write candidates", int(((proposal["write_risk"] == "high") & has_updates).sum())],
        ["Low-risk rows already confirmed/no further write", int(((proposal["write_risk"] == "low") & ~has_updates & (proposal["proposed_action"] != "accept_no_change")).sum())],
        ["Output CSV", str(CSV_OUT)],
        ["Output XLSX", str(XLSX_OUT)],
    ]
    for row in rows:
        ws.append(row)
    ws.append([])
    ws.append(["Action", "Rows"])
    for action, count in action_counts.items():
        ws.append([action, int(count)])
    ws.append([])
    ws.append(["Risk", "Rows"])
    for risk, count in risk_counts.items():
        ws.append([risk, int(count)])

    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="0F766E")
        cell.font = Font(color="FFFFFF", bold=True)
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 90
    ws.freeze_panes = "A2"


def main() -> None:
    df = pd.read_csv(INFILE, dtype=str).fillna("")
    proposal = build_proposal(df)
    CSV_OUT.parent.mkdir(parents=True, exist_ok=True)
    proposal.to_csv(CSV_OUT, index=False, encoding="utf-8")

    manual = proposal[proposal["proposed_action"] == "manual_review_required"]
    hide = proposal[proposal["proposed_action"].str.contains("hide|confirm_existing_hide", regex=True)]
    merge = proposal[proposal["proposed_action"].str.contains("confirm_broad_duplicate")]
    accept_cleanup = proposal[proposal["proposed_action"].str.contains("accept_cleanup")]
    accept_no_change = proposal[proposal["proposed_action"].str.contains("accept_no_change|accept_valid_but_duplicate_hidden")]
    has_updates = proposal["proposed_updates"].astype(str).ne("{}")
    safe_low = proposal[(proposal["write_risk"] == "low") & (proposal["proposed_action"] != "accept_no_change") & has_updates]

    wb = Workbook()
    write_summary(wb, proposal)
    write_sheet(wb, "Safe_Low_Risk", safe_low, "0F766E")
    write_sheet(wb, "Hide_Proposals", hide, "7C2D12")
    write_sheet(wb, "Merge_Proposals", merge, "581C87")
    write_sheet(wb, "Accept_Cleanup_Candidates", accept_cleanup, "1F4E79")
    write_sheet(wb, "Manual_Review", manual, "334155")
    write_sheet(wb, "Accept_No_Change", accept_no_change, "475569")
    write_sheet(wb, "All_Proposals", proposal, "111827")
    wb.save(XLSX_OUT)

    print(f"Wrote CSV: {CSV_OUT}")
    print(f"Wrote XLSX: {XLSX_OUT}")
    print("Action counts:")
    print(proposal["proposed_action"].value_counts().to_string())
    print("Risk counts:")
    print(proposal["write_risk"].value_counts().to_string())


if __name__ == "__main__":
    main()
